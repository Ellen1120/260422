"""
표준품 (Reference Standard) 조회 모듈
매칭 전략: 1) exact lowercase → 2) Excel 이름이 쿼리로 시작 → 3) stub

엑셀 파일 탐색 순서:
  1. 네트워크 경로 (\\\\file\\04. 품질본부\\...\\02.Working standard)
  2. 바로가기(.lnk) 해석 → 대상 경로 사용
  3. 접근 불가 시 → standards_db.py 옆 폴더의 .xlsx (fallback)
"""
from __future__ import annotations
import re
import subprocess
from pathlib import Path

_XLSX_NAME = "LIMS 도입 후 표준품 리스트_WS_최종본_250310.xlsx"
_NETWORK_XLSX = Path(r"\\file\04. 품질본부\3. 품질관리담당\1. 담당 공용\04. Standard & Column & Reagent\02.Working standard") / _XLSX_NAME
_LNK_PATH = Path(__file__).parent.parent / (_XLSX_NAME.replace(".xlsx", ".lnk"))
_FALLBACK_XLSX = Path(__file__).parent.parent / _XLSX_NAME

_DB: dict[str, list[dict]] = {}
_ALL_ENTRIES: list[dict] = []


def _resolve_lnk(lnk: Path) -> Path | None:
    """Windows 바로가기(.lnk)의 실제 대상 경로를 반환."""
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             f"(New-Object -ComObject WScript.Shell).CreateShortcut('{lnk}').TargetPath"],
            capture_output=True, text=True, timeout=5,
        )
        target = result.stdout.strip()
        if target:
            return Path(target)
    except Exception:
        pass
    return None


def _find_excel_path() -> Path | None:
    """사용할 엑셀 파일 경로를 반환. 우선순위: 네트워크 → .lnk 해석 → 로컬 fallback"""
    if _NETWORK_XLSX.exists():
        print(f"[standards_db] 네트워크 경로 사용: {_NETWORK_XLSX}")
        return _NETWORK_XLSX

    if _LNK_PATH.exists():
        resolved = _resolve_lnk(_LNK_PATH)
        if resolved and resolved.exists():
            print(f"[standards_db] 바로가기 경로 사용: {resolved}")
            return resolved
        print(f"[standards_db] 바로가기 대상 접근 불가: {resolved}")

    print(f"[standards_db] fallback 시도: {_FALLBACK_XLSX}")
    if _FALLBACK_XLSX.exists():
        return _FALLBACK_XLSX

    print("[standards_db] 표준품 엑셀 파일을 찾을 수 없습니다.")
    return None


def _load() -> tuple[dict[str, list[dict]], list[dict]]:
    import openpyxl

    excel_path = _find_excel_path()
    if excel_path is None:
        raise FileNotFoundError("표준품 엑셀 파일을 찾을 수 없습니다.")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    name_i = header.index("Name")
    ct_i   = header.index("Consumable Type")
    loc_i  = header.index("Location")

    db: dict[str, list[dict]] = {}
    seen_dedup: set[tuple] = set()
    all_entries: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_i]
        ct   = row[ct_i]
        loc  = row[loc_i]
        if not name:
            continue
        dedup_key = (str(name).strip(), str(ct or ""), str(loc or ""))
        if dedup_key in seen_dedup:
            continue
        seen_dedup.add(dedup_key)

        entry = {
            "name":            str(name).strip(),
            "consumable_type": str(ct or "").strip(),
            "location":        str(loc or "").strip(),
        }
        nk = entry["name"].lower()
        db.setdefault(nk, []).append(entry)
        all_entries.append(entry)

    wb.close()
    return db, all_entries


_ALIASES: dict[str, str] = {
    "scb4-impurity 1": "scb-4-imp-1",
    "sacubitril valsartan sodium hydrate": "sacubitril valsartan sodium",
}


def _find_matches(nm: str) -> list[dict]:
    """단일 이름에 대해 DB에서 매칭 엔트리 반환. 없으면 stub."""
    key = _ALIASES.get(nm.strip().lower(), nm.strip().lower())

    # 1. Exact match
    if key in _DB:
        return list(_DB[key])

    # 2. Prefix match
    prefix_matches: list[dict] = []
    seen_dk: set[tuple] = set()
    for entry in _ALL_ENTRIES:
        ek = entry["name"].lower()
        if ek == key or ek.startswith(key + " "):
            dk = (entry["name"], entry["consumable_type"], entry["location"])
            if dk not in seen_dk:
                seen_dk.add(dk)
                prefix_matches.append(entry)
    if prefix_matches:
        return prefix_matches

    # 3. Stub — STM에 명시된 표준품이지만 Excel에 등록되지 않은 경우
    return [{"name": nm.strip(), "consumable_type": "", "location": ""}]


def lookup(names: list[str]) -> list[dict]:
    """STM에서 추출한 표준품 이름 목록을 조회해 반환."""
    global _DB, _ALL_ENTRIES
    if not _DB:
        try:
            _DB, _ALL_ENTRIES = _load()
        except Exception as e:
            print(f"[standards_db] Excel 로드 실패: {e}")
            return [{"name": nm.strip(), "consumable_type": "", "location": ""} for nm in names]

    out: list[dict] = []
    seen: set[tuple] = set()
    for nm in names:
        for m in _find_matches(nm):
            dk = (m["name"], m["consumable_type"], m["location"])
            if dk not in seen:
                seen.add(dk)
                out.append(m)
    return out


def reload() -> None:
    """캐시를 초기화하고 엑셀을 다시 로드합니다 (서버 재시작 없이 갱신 가능)."""
    global _DB, _ALL_ENTRIES
    _DB = {}
    _ALL_ENTRIES = []
    try:
        _DB, _ALL_ENTRIES = _load()
        print(f"[standards_db] 재로드 완료: {len(_ALL_ENTRIES)}개 항목")
    except Exception as e:
        print(f"[standards_db] 재로드 실패: {e}")
