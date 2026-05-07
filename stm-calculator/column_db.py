"""컬럼 조회 모듈
LC01004-A1(4.0) Column List.xlsx LC 시트에서 STM 스펙 매칭 후 활성 컬럼 반환.

엑셀 파일 탐색 순서:
  1. 네트워크 경로 (\\\\file\\04. 품질본부\\3. 품질관리담당\\1. 담당 공용\\AI)
  2. 접근 불가 시 → column_db.py 옆 폴더의 .xlsx (fallback)
"""
from __future__ import annotations
import re
from pathlib import Path

_NETWORK_AI = Path(r"\\file\04. 품질본부\3. 품질관리담당\1. 담당 공용\AI")
_NETWORK_XLSX = _NETWORK_AI / "LC01004-A1(4.0) Column List.xlsx"
_FALLBACK_XLSX = Path(__file__).parent.parent / "LC01004-A1(4.0) Column List.xlsx"

_ENTRIES: list[dict] = []


def _find_excel_path() -> Path | None:
    """사용할 엑셀 파일 경로를 반환. 우선순위: 네트워크 → 로컬 fallback"""
    if _NETWORK_XLSX.exists():
        print(f"[column_db] 네트워크 경로 사용: {_NETWORK_XLSX}")
        return _NETWORK_XLSX

    print(f"[column_db] 네트워크 접근 불가, fallback 시도: {_FALLBACK_XLSX}")
    if _FALLBACK_XLSX.exists():
        return _FALLBACK_XLSX

    print("[column_db] 컬럼 엑셀 파일을 찾을 수 없습니다.")
    return None


def _parse_excel_spec(spec: str) -> dict:
    """'150*4.6mm, 3um' → {length:150, id:4.6, particle:3}"""
    nums = re.findall(r'\d+(?:\.\d+)?', spec)
    if len(nums) >= 3:
        a, b = float(nums[0]), float(nums[1])
        return {"length": max(a, b), "id": min(a, b), "particle": float(nums[2])}
    if len(nums) == 2:
        return {"length": float(nums[0]), "id": None, "particle": float(nums[1])}
    return {}


def _parse_stm_col(stm_spec: str) -> tuple[dict, set[str]]:
    """'YMC Triart (C18), 4.6 x 150 mm, 3 µm' → (spec_dict, name_keywords)"""
    particle_m = re.search(r'(\d+(?:\.\d+)?)\s*[µu]m', stm_spec, re.IGNORECASE)
    size_m = re.search(
        r'(\d+(?:\.\d+)?)\s*(mm|cm)?\s*[x×]\s*(\d+(?:\.\d+)?)\s*(mm|cm)?',
        stm_spec, re.IGNORECASE,
    )

    id_ = length = None
    if size_m:
        a = float(size_m.group(1)) * (10 if (size_m.group(2) or '').lower() == 'cm' else 1)
        b = float(size_m.group(3)) * (10 if (size_m.group(4) or '').lower() == 'cm' else 1)
        id_, length = min(a, b), max(a, b)
    particle = float(particle_m.group(1)) if particle_m else None

    name_part = stm_spec.split(',')[0]
    tokens = re.sub(r'[^a-z0-9]', ' ', name_part.lower()).split()
    stopwords = {'mm', 'um', 'ml', 'x', 'uv', 'nm', 'min', 'and', 'the', 'of', 'in', 'm', 'c18', 'c8', 'c4', 'c1'}
    keywords = {t for t in tokens if len(t) >= 2 and re.search(r'[a-z]', t) and t not in stopwords}

    return {"length": length, "id": id_, "particle": particle}, keywords


def _spec_matches(excel: dict, stm: dict) -> bool:
    for key in ("length", "id", "particle"):
        ev, sv = excel.get(key), stm.get(key)
        if ev is not None and sv is not None and ev != sv:
            return False
    return True


def _name_matches(excel_name_lower: str, stm_keywords: set[str]) -> bool:
    if not stm_keywords:
        return False
    return all(k in excel_name_lower for k in stm_keywords)


def _load() -> list[dict]:
    import openpyxl

    excel_path = _find_excel_path()
    if excel_path is None:
        raise FileNotFoundError("컬럼 엑셀 파일을 찾을 수 없습니다.")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb['LC']
    header = [str(c.value).strip().replace('\n', ' ') if c.value else '' for c in ws[1]]
    no_i    = header.index('Column No.')
    name_i  = header.index('Name')
    spec_i  = header.index('Spec')
    loc_i   = header.index('Storage Location')
    disc_i  = header.index('Discarded Date')
    test_i  = header.index('Test Item')
    init_i  = header.index('Initiation Date')

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[disc_i]:
            continue
        no   = str(row[no_i]   or '').strip()
        name = str(row[name_i] or '').strip()
        spec = str(row[spec_i] or '').strip()
        loc  = str(row[loc_i]  or '').strip()
        if not no or not name:
            continue
        test_item = str(row[test_i] or '').strip()
        initiated = bool(row[init_i])
        entries.append({
            "column_no":    no,
            "name":         name,
            "spec":         spec,
            "location":     loc,
            "test_item":    test_item,
            "remark":       "" if initiated else "미개봉",
            "_name_lower":  name.lower(),
            "_spec_parsed": _parse_excel_spec(spec),
        })
    wb.close()
    return entries


def lookup(stm_col_specs: list[str]) -> list[dict]:
    """STM 컬럼 스펙 문자열 목록으로 매칭되는 활성 컬럼 목록 반환."""
    global _ENTRIES
    if not _ENTRIES:
        try:
            _ENTRIES = _load()
        except Exception as e:
            print(f"[column_db] Excel 로드 실패: {e}")
            return []

    results: list[dict] = []
    seen: set[str] = set()

    for stm_spec in stm_col_specs:
        if not stm_spec:
            continue
        stm_parsed, stm_keywords = _parse_stm_col(stm_spec)
        for entry in _ENTRIES:
            no = entry["column_no"]
            if no in seen:
                continue
            if (_name_matches(entry["_name_lower"], stm_keywords) and
                    _spec_matches(entry["_spec_parsed"], stm_parsed)):
                seen.add(no)
                results.append({
                    "column_no": no,
                    "name":      entry["name"],
                    "spec":      entry["spec"],
                    "location":  entry["location"],
                    "test_item": entry["test_item"],
                    "remark":    entry["remark"],
                })

    return results


def reload() -> None:
    """캐시를 초기화하고 엑셀을 다시 로드합니다 (서버 재시작 없이 갱신 가능)."""
    global _ENTRIES
    _ENTRIES = []
    try:
        _ENTRIES = _load()
        print(f"[column_db] 재로드 완료: {len(_ENTRIES)}개 컬럼")
    except Exception as e:
        print(f"[column_db] 재로드 실패: {e}")
